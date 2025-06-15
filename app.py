import pyautogui
import time
import webbrowser
from urllib.parse import quote
import openpyxl
import tkinter as tk
from tkinter import ttk
import threading

# Variável global para controle do envio
envio_ativo = threading.Event()

# Lista global com os tipos de mensagem
tipos_mensagem = [
    "Todos", "prospects", "pos", "em negociacao", "parceiros", "contadores","em treinamento",
    "smart", "evolutiplay", "tef", "nfe", "não usa", "pouco suporte", "usa bem", "mei"
]

def enviar_mensagem_com_enter(cliente, mensagem):
    try:
        telefone_formatado = cliente["telefone"]
        link_mensagem_whatsapp = f'https://web.whatsapp.com/send?phone={telefone_formatado}&text={quote(mensagem)}'
        webbrowser.open(link_mensagem_whatsapp)
        time.sleep(20)
        pyautogui.press('enter')
        time.sleep(5)
    except Exception as e:
        print(f"Erro ao enviar para {cliente['nome']}: {e}")

def criar_mensagem(cliente):
    grupo = cliente["grupo"].strip().lower()
    nome = cliente["nome"]
    empresa = cliente["empresa"]
    inicio = cliente["inicio"] if cliente["inicio"] else "Indefinido"

#MONTANDO A MENSAGEM PERSONALIZADA PARA TIPO\GRUPO DE CLIENTES\PROSPECTS
def criar_mensagem(cliente):
    grupo = cliente["grupo"].strip().lower()
    nome = cliente["nome"]
    empresa = cliente["empresa"]
    inicio = cliente["inicio"] if cliente["inicio"] else "Indefinido"

    mensagens = {
        "em treinamento": (
            f"{inicio}, {nome}, da empresa {empresa}, "
            "Como está o uso do *Sistema EvoluTI*? Tudo certo? Algo a relatar? "
            "Prestamos suporte técnico rápido e ativo. Qualquer coisa, pode entrar em contato comigo, o Glaucio, "
            "ou com nosso Suporte no *(55) 9119 4370* (Bruna). "
            "Caso não queira mais receber essa forma de contato, envie *SAIR*."
        ),
        "prospects": (
            f"{inicio}, {nome}, da empresa {empresa}, "
            "Nós da *GDI Informática* trabalhamos com um *Sistema de Gestão: Simples e Prático*. "
            "Montamos a mensalidade baseada nas ferramentas que realmente for utilizar. "
            "Valores a partir de R$ 69,90. "
            "Se não quiser mais receber informações sobre nossos serviços e produtos, envie *SAIR*."
        ),
        "pos": (
            f"{inicio}, {nome}, da empresa {empresa}, "
            "Como está o uso do aplicativo na maquininha de cartões? Tudo certo? Algo a relatar? "
            "Entre em contato para mais informações no (54) 9 9104 1029. "
            "Prestamos suporte técnico rápido e ativo. "
            "Caso não queira mais receber informações, envie *SAIR*."
        ),
        "em negociacao": (
            f"{inicio}, {nome}, da empresa {empresa}, "
            "Nós da *GDI Informática* trabalhamos com um *Sistema de Gestão: Simples e Prático*. "
            "Já tínhamos lhe visitado anteriormente, e agora temos *Novidades* e *Promoções* exclusivas para você. "
            "Entre em contato comigo, o Glaucio! "
            "Caso não queira mais receber esse tipo de contato, envie *SAIR*."
        ),
        "parceiros": (
            f"{inicio}, {nome}, da empresa {empresa}, "
            "Nós da *GDI Informática* reafirmamos nossa parceria e temos *Novidades* e *Promoções* "
            "exclusivas para seus clientes nesse novo ciclo. "
            "Entre em contato comigo, o Glaucio! "
            "Caso não queira mais receber informações, envie *SAIR*."
        ),
        "contadores": (
            f"{inicio}, {nome}, da empresa {empresa}, "
            "Visitei seu escritório pessoalmente, e estamos entrando em contato novamente para apresentar uma "
            "solução prática e simples para seu cliente com relação ao *Sistema de Gestão*. "
            "Entre em contato comigo, o Glaucio! "
            "Caso não queira mais receber informações, envie *SAIR*."
        ),
        "smart": (
            f"{inicio}, {nome}, da empresa {empresa}, "
            "Como está o uso do *Sistema EvoluTI*, tudo certo? Sabemos que utiliza a maquininha Smart para emissão do cupom. "
            "Às vezes acontecem alguns problemas, mas agora temos a opção de ligação direta nas máquinas do Banrisul ou Sicredi. "
            "Se quiser trocar para essa nova forma, fale comigo, o Glaucio, ou com nosso suporte no *(55) 9119 4370* (Bruna). "
            "Caso não queira mais receber essa forma de contato, envie *SAIR*."
        ),
        "evolutiplay": (
            f"{inicio}, {nome}, da empresa {empresa}, "
            "Como está o uso do *Sistema EvoluTI*, tudo certo? "
            "Temos a solução para ligação direta nas máquinas de cartão SMART, utilizando Banrisul, "
            "atendendo às exigências fiscais atuais. "
            "Entre em contato comigo, o Glaucio, ou com nosso suporte no *(55) 9119 4370* (Bruna). "
            "Caso não queira mais receber essa forma de contato, envie *SAIR*."
        ),
        "tef": (
            f"{inicio}, {nome}, da empresa {empresa}, "
            "Como está o uso do *Sistema EvoluTI*, tudo certo? Sabemos que utiliza o TEF na emissão de cupom eletrônico. "
            "Às vezes surgem problemas, mas agora temos a opção de ligação direta nas máquinas de cartão (SMART, Banrisul ou Sicredi). "
            "Se quiser trocar para essa nova forma, fale comigo, o Glaucio, ou com nosso suporte no *(55) 9119 4370* (Bruna). "
            "Caso não queira mais receber essa forma de contato, envie *SAIR*."
        ),
        "nfe": (
            f"{inicio}, {nome}, da empresa {empresa}, "
            "Como está o uso do *Sistema EvoluTI*, tudo certo? "
            "Sabemos que utiliza principalmente a emissão de nota eletrônica. "
            "Estamos sempre à disposição para ajudar. "
            "Fale comigo, o Glaucio, ou com nosso suporte no *(55) 9119 4370* (Bruna). "
            "Caso não queira mais receber esse tipo de contato, envie *SAIR*."
        ),
        "pouco suporte": (
            f"{inicio}, {nome}, da empresa {empresa}, "
            "Como está o uso do *Sistema EvoluTI*? "
            "Sabemos que dificilmente precisa de ajuda. "
            "Mas estamos prontos para ajudar, caso precisar de algo. "
            "Se não quiser mais receber informações, envie *SAIR*."
        ),    
        "não usa": (
            f"{inicio}, {nome}, da empresa {empresa}, "
            "Como está o uso do *Sistema EvoluTI*? "
            "Sabemos que utiliza poucas opções e ferramentas. "
            "Estamos prontos para ajudar ou implementar novos processos conforme necessário. "
            "Se não quiser mais receber informações, envie *SAIR*."
        ),
        "usa bem": (
            f"{inicio}, {nome}, da empresa {empresa}, "
            "Como está o uso do *Sistema EvoluTI*, tudo certo? "
            "Além de todas as opções que já utiliza, estamos sempre à disposição para ajudar. "
            "Se não quiser mais receber informações, envie *SAIR*."
        ),
        "mei": (
            f"{inicio}, {nome}, da empresa {empresa}, "
            "Como está o uso do *Sistema EvoluTI*, tudo certo? "
            "Além das opções que já utiliza, temos suporte para a parte fiscal, incluindo emissão de NF-e, cupom eletrônico e TEF. "
            "Se não quiser mais receber informações, envie *SAIR*."
        ),
    }

    return mensagens.get(grupo, f"Olá {nome}, da empresa {empresa}, Somos da GDI INFORMATICA, um Sistema de Gestão: *Simples e Prático!*")

def ler_dados_planilha(pagina_selecionada, grupo_selecionado):
    try:
        workbook = openpyxl.load_workbook('Mala_Whats.xlsx')
        aba = workbook[pagina_selecionada]
        clientes = []
        for linha in aba.iter_rows(min_row=2, values_only=True):
            if linha[2]:
                cliente = {
                    "empresa": linha[0],
                    "nome": linha[1],
                    "telefone": linha[2],
                    "inicio": linha[3] if linha[3] else "Indefinido",
                    "grupo": linha[4].strip().lower() if linha[4] else "indefinido"
                }
                if grupo_selecionado.lower() == "todos" or cliente["grupo"] == grupo_selecionado.lower():
                    clientes.append(cliente)
        return clientes
    except Exception as e:
        print(f"Erro ao ler a planilha: {e}")
        return []

# === Interface Gráfica ===
janela = tk.Tk()
janela.title('Envio Automático de Mensagens WhatsApp')
janela.attributes('-fullscreen', True)

# Controle de tela cheia
tela_cheia = True

def alternar_tela_cheia():
    global tela_cheia
    tela_cheia = not tela_cheia
    janela.attributes('-fullscreen', tela_cheia)

def sair_tela_cheia(event=None):
    global tela_cheia
    tela_cheia = False
    janela.attributes('-fullscreen', False)

# Dados do Excel
abas = openpyxl.load_workbook('Mala_Whats.xlsx').sheetnames
pagina_var = tk.StringVar(value=abas[0])
grupo_var = tk.StringVar(value="Todos")

# Frame de seleção
frame_selecao = tk.Frame(janela, pady=10)
frame_selecao.pack()

tk.Label(frame_selecao, text="Página:", font=('Arial', 12, 'bold')).grid(row=0, column=0, padx=10, pady=5, sticky="w")
tk.OptionMenu(frame_selecao, pagina_var, *abas).grid(row=0, column=1, padx=10, pady=5)

tk.Label(frame_selecao, text="Tipo de Mensagem:", font=('Arial', 12, 'bold')).grid(row=0, column=2, padx=10, pady=5, sticky="w")
tk.OptionMenu(frame_selecao, grupo_var, *tipos_mensagem).grid(row=0, column=3, padx=10, pady=5)

# Frame da Treeview
frame = ttk.Frame(janela, padding=10)
frame.pack(fill=tk.BOTH, expand=True)

tree = ttk.Treeview(frame, columns=("Empresa", "Telefone", "Mensagem"), show="headings")
tree.heading("Empresa", text="Empresa")
tree.heading("Telefone", text="Telefone")
tree.heading("Mensagem", text="Mensagem")

tree.column("Empresa", width=200, anchor="w", stretch=False)
tree.column("Telefone", width=150, anchor="center", stretch=False)
tree.column("Mensagem", width=1200, anchor="w", stretch=True)

tree.pack(fill=tk.BOTH, expand=True)

def ajustar_colunas(event):
    largura_total = tree.winfo_width()
    tree.column("Empresa", width=int(largura_total * 0.15))
    tree.column("Telefone", width=int(largura_total * 0.15))
    tree.column("Mensagem", width=int(largura_total * 0.70))

tree.bind("<Configure>", ajustar_colunas)

# Contagem

label_contagem = tk.Label(janela, text="Total de Mensagens: 0", font=('Arial', 12, 'bold'))
label_contagem.pack(pady=5)

def atualizar_contagem():
    total = len(tree.get_children())
    label_contagem.config(text=f"Total de Mensagens: {total}")

# Carregar dados
def carregar_dados():
    for row in tree.get_children():
        tree.delete(row)
    try:
        workbook = openpyxl.load_workbook('Mala_Whats.xlsx')
        aba = workbook[pagina_var.get()]
        tipo_selecionado = grupo_var.get().strip().lower()
        for linha in aba.iter_rows(min_row=2, values_only=True):
            if linha[2]:
                tipo_mensagem = linha[4].strip().lower() if linha[4] else "indefinido"
                if tipo_selecionado == "todos" or tipo_mensagem == tipo_selecionado:
                    mensagem = criar_mensagem({
                        "empresa": linha[0],
                        "nome": linha[1],
                        "telefone": linha[2],
                        "inicio": linha[3] if linha[3] else "Indefinido",
                        "grupo": linha[4] if linha[4] else "indefinido"
                    })
                    tree.insert("", tk.END, values=(linha[1], linha[2], mensagem))
        atualizar_contagem()
    except Exception as e:
        print(f"Erro ao carregar dados: {e}")

# Editar mensagem
def editar_mensagem():
    selected_item = tree.selection()
    if selected_item:
        item = tree.item(selected_item)
        top = tk.Toplevel(janela)
        top.title("Editar Mensagem")
        tk.Label(top, text="Mensagem:").pack()
        texto = tk.Text(top, height=15, width=120)
        texto.insert(tk.END, item['values'][2])
        texto.pack()
        def salvar():
            tree.item(selected_item, values=(item['values'][0], item['values'][1], texto.get("1.0", tk.END).strip()))
            top.destroy()
        tk.Button(top, text="Salvar", command=salvar).pack()

# Envio de mensagens
def iniciar_envio_thread():
    envio_ativo.set()
    dados_envio = []

    for item in tree.get_children():
        nome, telefone, mensagem = tree.item(item, "values")
        dados_envio.append({
            "nome": nome,
            "telefone": telefone,
            "mensagem": mensagem
        })

    def enviar_mensagens():
        for cliente in dados_envio:
            if not envio_ativo.is_set():
                print("Envio cancelado.")
                break
            enviar_mensagem_com_enter(cliente, cliente["mensagem"])
            time.sleep(105)  # Tempo de espera entre mensagens
        print("Envio concluído!")

    threading.Thread(target=enviar_mensagens, daemon=True).start()

def cancelar_envio():
    envio_ativo.clear()
    print("Envio cancelado!")

def deletar_mensagem():
    selected_item = tree.selection()
    if selected_item:
        tree.delete(selected_item)
        atualizar_contagem()

# Frame dos Botões
frame_botoes = tk.Frame(janela, pady=10)
frame_botoes.pack()

tk.Button(frame_botoes, text="Carregar Dados", command=carregar_dados, bg="blue", fg="white", width=15).grid(row=0, column=0, padx=8, pady=5)
tk.Button(frame_botoes, text="Editar Mensagem", command=editar_mensagem, bg="gray", fg="white", width=15).grid(row=0, column=1, padx=8, pady=5)
tk.Button(frame_botoes, text="Iniciar Envio", command=iniciar_envio_thread, bg="green", fg="white", width=15).grid(row=0, column=2, padx=8, pady=5)
tk.Button(frame_botoes, text="Cancelar Envio", command=cancelar_envio, bg="red", fg="white", width=15).grid(row=0, column=3, padx=8, pady=5)
tk.Button(frame_botoes, text="Excluir Mensagem", command=deletar_mensagem, bg="orange", fg="black", width=15).grid(row=0, column=4, padx=8, pady=5)
tk.Button(frame_botoes, text="Fechar", command=janela.destroy, bg="black", fg="white", width=15).grid(row=0, column=5, padx=8, pady=5)

# 🔥 Atalho ESC para sair do modo tela cheia
janela.bind("<Escape>", sair_tela_cheia)

janela.mainloop()
